from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, abort
import csv
import os
from datetime import datetime, time, timedelta
import pytz
import pandas as pd  # For exporting to Excel
import requests  # For fetching time from external API
from functools import wraps
from dotenv import load_dotenv
from io import BytesIO
from openpyxl.styles import PatternFill
import logging
import shutil  # Added for backup
import uuid
import json

def purge_duplicate_actions():
    """
    Purges duplicate Time-In and Time-Out actions in the log.csv file,
    keeping only the first occurrence for each user per date.
    """
    if not os.path.isfile(LOG_FILE):
        app.logger.warning("Log file does not exist. No duplicates to purge.")
        return False, "Log file does not exist. No duplicates to purge."

    try:
        # Backup the original log file before making changes
        backup_file = LOG_FILE + ".backup"
        shutil.copy(LOG_FILE, backup_file)
        app.logger.info(f"Backup of log file created at {backup_file}.")

        # Read the log CSV into a DataFrame
        df = pd.read_csv(LOG_FILE, encoding='utf-8')

        # Sort by ID ascending to keep the first occurrence
        df_sorted = df.sort_values(by='ID', ascending=True)

        # Drop duplicates based on Name, Date, and Action, keeping the first occurrence
        df_cleaned = df_sorted.drop_duplicates(subset=['Name', 'Date', 'Action'], keep='first')

        # Check if any duplicates were removed
        duplicates_removed = len(df_sorted) - len(df_cleaned)
        if duplicates_removed == 0:
            app.logger.info("No duplicate actions found in the log file.")
            return False, "No duplicate actions found in the log file."

        # Save the cleaned DataFrame back to the CSV
        df_cleaned.to_csv(LOG_FILE, index=False, quoting=csv.QUOTE_ALL)
        app.logger.info(f"Purged {duplicates_removed} duplicate actions from the log file.")
        return True, f"Purged {duplicates_removed} duplicate actions from the log file."
    except Exception as e:
        app.logger.error(f"Error purging duplicate actions: {e}")
        return False, f"Error purging duplicate actions: {e}"

# Load environment variables from a .env file if present
load_dotenv()

# Configure logging
logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(name)s %(threadName)s : %(message)s'
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EMPLOYEES_FILE = os.path.join(BASE_DIR, 'employees.csv')

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your_default_secret_key')  # Use environment variable for secret key

# Session Configuration for Enhanced Security
app.config.update(
    SESSION_COOKIE_SECURE=True,      # Ensure cookies are sent over HTTPS
    SESSION_COOKIE_HTTPONLY=True,    # Prevent JavaScript from accessing the cookie
    SESSION_COOKIE_SAMESITE='Lax'    # Adjust as per your requirements
)

# Define Pakistan time zone
LOCAL_TIME_ZONE = pytz.timezone('Asia/Karachi')  # Use Pakistan time zone

# Configurable shift times
SHIFT_START = time(6, 0)       # 6:00 AM
SHIFT_END = time(11, 59)       # 11:59 AM
EXPECTED_TIME_IN = time(8, 0)  # 8:00 AM

# PM Shift times
PM_SHIFT_START = time(18, 0)   # 6:00 PM
PM_SHIFT_END = time(23, 59)    # 11:59 PM
PM_EXPECTED_TIME_IN = time(20, 0)  # 8:00 PM

# Path to the m_credential CSV and log CSV
m_credential_FILE = os.path.join(BASE_DIR, 'm_credential.csv')
LOG_FILE = os.path.join(BASE_DIR, 'log.csv')

# Define time limits for actions
TIME_LIMITS = {
    "Recite Sutra": 30,
    "Toilet": 20,
    "Smoke": 20,
    "BREAK1": 45,
    "BREAK2": 45,
}


def get_employee_list():
    """Reads the employee list from employees.csv and returns a list of dictionaries."""
    employee_list = []
    if os.path.isfile(EMPLOYEES_FILE):
        try:
            with open(EMPLOYEES_FILE, 'r', newline='', encoding='utf-8') as csvfile:
                csvreader = csv.DictReader(csvfile)
                for row in csvreader:
                    # Ensure that 'ID' is treated as a string
                    employee_list.append({'ID': row['ID'].zfill(4), 'Name': row['Name']})
        except Exception as e:
            app.logger.error(f"Error reading employee list: {e}")
    else:
        app.logger.warning("Employee list file does not exist.")
    return employee_list

def get_keys():
    """Read the master key and all sub-keys from the CSV file."""
    if os.path.isfile(m_credential_FILE):
        with open(m_credential_FILE, 'r', newline='', encoding='utf-8') as csvfile:
            csvreader = csv.reader(csvfile)
            headers = next(csvreader, None)  # Skip header
            for row in csvreader:
                if row:
                    master_key = row[0].strip()
                    sub_keys = [key.strip() for key in row[1:]]  # Get all the sub-keys
                    return master_key, sub_keys
    return None, []

def set_keys(master_key, sub_keys):
    """Write the master key and sub-keys to the CSV file."""
    with open(m_credential_FILE, 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile)
        headers = ['master_key'] + [f'sub_key{i+1}' for i in range(len(sub_keys))]
        csvwriter.writerow(headers)
        csvwriter.writerow([master_key] + sub_keys)

def get_pakistan_time():
    """Fetch the current time in Pakistan timezone from an external API."""
    try:
        response = requests.get('http://worldtimeapi.org/api/timezone/Asia/Karachi', timeout=5)
        if response.status_code == 200:
            data = response.json()
            datetime_str = data['datetime']  # ISO 8601 format
            timestamp = datetime.fromisoformat(datetime_str[:-1])  # Remove the 'Z' at the end
            return LOCAL_TIME_ZONE.localize(timestamp)
        else:
            app.logger.warning("Error fetching time from API, using local time.")
            return datetime.now(LOCAL_TIME_ZONE)
    except Exception as e:
        app.logger.error(f"Exception occurred while fetching time: {e}")
        return datetime.now(LOCAL_TIME_ZONE)

def login_required(f):
    """Decorator to ensure the user is authenticated."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('authenticated'):
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to ensure the user has admin privileges."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('report'))
        return f(*args, **kwargs)
    return decorated_function

def handle_halfday_time_in(employee_id, name, group, timestamp, date_str, time_str):
    """
    Handles the Halfday Time-In action by recording it without enforcing schedule.
    """
    # Initialize log ID
    new_id = get_next_log_id()

    # Log the data
    data = {
        'ID': new_id,
        'Employee ID': int(employee_id),
        'Name': name,
        'Group': group.upper(),
        'Action': 'Halfday_Time_In',
        'Date': date_str,
        'Start Time': time_str,
        'End Time': '',
        'Time Consumed': '',
        'Shift': 'Halfday',
        'Lateness Duration': '',
        'Status': 'Halfday Time-In'
    }

    # Append to log.csv
    append_to_log_file(data)

    flash(f"Halfday Time-In recorded for {name} on {date_str} at {time_str}.", 'info')
    return redirect(url_for('index'))

def handle_halfday_time_out(employee_id, name, group, timestamp, date_str, time_str):
    """
    Handles the Halfday Time-Out action by updating the corresponding Halfday Time-In entry.
    """
    if os.path.isfile(LOG_FILE):
        try:
            df = pd.read_csv(LOG_FILE, encoding='utf-8')
            # Fill NaN values
            df['Action'] = df['Action'].fillna('')
            df['End Time'] = df['End Time'].fillna('')
            df['Date'] = df['Date'].fillna('')
            df['Start Time'] = df['Start Time'].fillna('')
            # Find the last Halfday Time-In entry for this user and date without End Time
            mask = (
                (df['Employee ID'] == int(employee_id)) &
                (df['Date'].str.strip() == date_str) &
                (df['Action'].str.lower() == 'halfday_time_in') &
                (df['End Time'].str.strip() == '')
            )
            if not mask.any():
                flash('Cannot Halfday Time-Out without Halfday Time-In first.', 'warning')
                return redirect(url_for('index'))
            else:
                # Get index of the entry
                idx = df[mask].index[-1]
                # Update the entry
                end_time_str = time_str
                start_time_str = df.loc[idx, 'Start Time']
                if not start_time_str:
                    flash('Start Time is missing for Halfday Time-In. Cannot record Halfday Time-Out.', 'danger')
                    return redirect(url_for('index'))
                # Parse times
                start_time = datetime.strptime(start_time_str, '%H:%M:%S')
                end_time = datetime.strptime(end_time_str, '%H:%M:%S')
                # If end_time < start_time, it means the end time is on the next day
                if end_time < start_time:
                    end_time += timedelta(days=1)
                duration_td = end_time - start_time
                duration_seconds = duration_td.total_seconds()

                # Extract hours, minutes, and seconds
                duration_hours = int(duration_seconds // 3600)
                duration_remaining_seconds = int(duration_seconds % 3600)
                duration_minutes = duration_remaining_seconds // 60
                duration_secs = duration_remaining_seconds % 60

                # Format Time Consumed
                duration_str_parts = []
                if duration_hours > 0:
                    duration_str_parts.append(f"{duration_hours} hrs")
                if duration_minutes > 0:
                    duration_str_parts.append(f"{duration_minutes} mins")
                if duration_secs > 0:
                    duration_str_parts.append(f"{duration_secs} secs")
                duration_str = ' & '.join(duration_str_parts) if duration_str_parts else '0 secs'

                # Update the Action to combine Halfday_Time_In and Halfday_Time_Out
                df.loc[idx, 'Action'] = 'Halfday_Time_In/Halfday_Time_Out'
                df.loc[idx, 'End Time'] = end_time_str
                df.loc[idx, 'Time Consumed'] = duration_str
                df.loc[idx, 'Status'] = 'Halfday Time-Out'

                # Save the DataFrame back to CSV
                df.to_csv(LOG_FILE, index=False, quoting=csv.QUOTE_ALL)
                app.logger.info(f"Halfday Time-Out recorded and combined for {name} on {date_str} at {end_time_str}.")
                flash(f"Halfday Time-Out recorded and combined for {name} on {date_str} at {end_time_str}.", 'info')
                return redirect(url_for('index'))
        except Exception as e:
            app.logger.error(f"Error processing Halfday Time-Out: {e}")
            flash('Failed to record Halfday Time-Out. Please try again.', 'danger')
            return redirect(url_for('index'))
    else:
        flash('Cannot Halfday Time-Out without Halfday Time-In first.', 'warning')
        return redirect(url_for('index'))

def get_next_log_id():
    """Retrieves the next available log ID."""
    if os.path.isfile(LOG_FILE):
        try:
            with open(LOG_FILE, 'r', newline='', encoding='utf-8') as csvfile:
                csvreader = csv.reader(csvfile)
                next(csvreader, None)  # Skip header
                last_id = 0
                for row in csvreader:
                    if row and row[0].isdigit():
                        last_id = int(row[0])
                return last_id + 1
        except Exception as e:
            app.logger.error(f"Error reading log file for next ID: {e}")
            return 1
    else:
        return 1

def append_to_log_file(data):
    """Appends a single record to the log.csv file."""
    file_exists = os.path.isfile(LOG_FILE)
    try:
        with open(LOG_FILE, 'a', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['ID', 'Employee ID', 'Name', 'Group', 'Action', 'Date',
                          'Start Time', 'End Time', 'Time Consumed', 'Shift',
                          'Lateness Duration', 'Status']
            csvwriter = csv.DictWriter(csvfile, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
            if not file_exists:
                csvwriter.writeheader()
            csvwriter.writerow(data)
    except Exception as e:
        app.logger.error(f"Error writing to log file: {e}")
        flash('Failed to record action. Please try again.', 'danger')


@app.route('/attendance', methods=['GET'])
def index():
    employee_list = get_employee_list()
    return render_template('index.html', employee_list=employee_list)

@app.route('/attendance/submit', methods=['POST'])
def submit():
    employee_id = request.form.get('employee_id', '').strip()
    group = request.form.get('group', '').strip().lower()
    action = request.form.get('action', '').strip()

    if not employee_id or not group or not action:
        flash('Invalid input data.', 'danger')
        return redirect(url_for('index'))

    # Validate the employee ID and get the name
    employee_list = get_employee_list()
    employee_dict = {emp['ID']: emp['Name'] for emp in employee_list}

    if employee_id not in employee_dict:
        flash('Invalid employee selected.', 'danger')
        return redirect(url_for('index'))

    name = employee_dict[employee_id]

    timestamp = get_pakistan_time()
    date_str = timestamp.strftime('%Y-%m-%d')
    time_str = timestamp.strftime('%H:%M:%S')

    # Define actions that can have duplicates
    ALLOW_DUPLICATES_ACTIONS = [action.strip().lower() for action in ['halfday_time_in', 'halfday_time_out'] + list(TIME_LIMITS.keys())]

    # Duplicate Action Check
    if action.lower() not in ALLOW_DUPLICATES_ACTIONS:
        if os.path.isfile(LOG_FILE):
            try:
                df = pd.read_csv(LOG_FILE, encoding='utf-8')
                # Fill NaN values
                df['Action'] = df['Action'].fillna('')
                df['Date'] = df['Date'].fillna('')
                df['Employee ID'] = df['Employee ID'].fillna('')

                # Check if the action already exists for the user on the same date
                mask = (
                    (df['Employee ID'] == int(employee_id)) &
                    (df['Date'].str.strip() == date_str) &
                    (df['Action'].str.lower() == action.lower())
                )
                if mask.any():
                    flash(f"You have already performed '{action}' today.", 'warning')
                    return redirect(url_for('index'))
            except Exception as e:
                app.logger.error(f"Error checking for duplicate action: {e}")
                flash('Failed to check for duplicate actions.', 'danger')
                return redirect(url_for('index'))


    BYPASS_TIME_IN_ACTIONS = ['halfday_time_in', 'halfday_time_out']

    if action.lower() not in ['time_in'] + BYPASS_TIME_IN_ACTIONS:
        # Check if user has a Time-In entry without End Time
        if os.path.isfile(LOG_FILE):
            try:
                df = pd.read_csv(LOG_FILE, encoding='utf-8')
                # Fill NaN values
                df['Action'] = df['Action'].fillna('')
                df['End Time'] = df['End Time'].fillna('')
                df['Date'] = df['Date'].fillna('')
                # Find if there's a Time-In for today without an End Time
                mask = (
                    (df['Employee ID'] == int(employee_id)) &
                    (df['Date'].str.strip() == date_str) &
                    (df['Action'].str.lower() == 'time_in') &
                    (df['End Time'].str.strip() == '')
                )
                if not mask.any():
                    flash('You must Time-In before performing other actions.', 'warning')
                    return redirect(url_for('index'))
            except Exception as e:
                app.logger.error(f"Error checking Time-In status: {e}")
                flash('Failed to check Time-In status.', 'danger')
                return redirect(url_for('index'))
        else:
            flash('You must Time-In before performing other actions.', 'warning')
            return redirect(url_for('index'))

    if action.lower() == 'time_in':
        # Existing Time-In logic
        current_time = timestamp.time()

        # Initialize expected_time_am and expected_time_pm based on group
        special_expected_time_am = EXPECTED_TIME_IN
        special_expected_time_pm = PM_EXPECTED_TIME_IN

        # Define group-specific shifts
        if group == 'hr':
            if current_time < time(10, 0):
                special_expected_time_am = time(8, 0)
                special_expected_time_pm = None  # No PM shift for HR
            else:
                special_expected_time_am = time(12, 0)
                special_expected_time_pm = None
        elif group in {'mqm', 'mkm', 'trainer'}:
            special_expected_time_am = time(8, 45)  # Day shift
            special_expected_time_pm = time(20, 45)  # Night shift
        elif group == 'office boy':
            special_expected_time_am = time(9, 0)
            special_expected_time_pm = time(21, 0)
        elif group in {'mdm', 'mbm', 'group leader', 'team leader'}:
            special_expected_time_am = time(8, 15)
            special_expected_time_pm = time(20, 15)
        elif group == 'admin':
            special_expected_time_am = time(11, 0)
            special_expected_time_pm = time(23, 0)

        # Determine shift and expected time
        if group == 'hr':
            if time(6, 0) <= current_time < time(10, 0):
                expected_time = datetime.combine(timestamp.date(), special_expected_time_am)
                shift = 'AM Shift'
            elif time(10, 0) <= current_time < time(18, 0):
                expected_time = datetime.combine(timestamp.date(), special_expected_time_am)
                shift = 'Midday Shift'
            elif PM_SHIFT_START <= current_time <= PM_SHIFT_END:
                expected_time = None
                shift = 'No PM Shift'
            elif current_time < SHIFT_START:
                if special_expected_time_pm:
                    expected_time = datetime.combine(timestamp.date() - timedelta(days=1), special_expected_time_pm)
                    shift = 'PM Shift (after midnight)'
                else:
                    expected_time = None
                    shift = 'No PM Shift'
            else:
                expected_time = None
                shift = 'Unknown'
        else:
            if SHIFT_START <= current_time <= SHIFT_END:
                expected_time = datetime.combine(timestamp.date(), special_expected_time_am)
                shift = 'AM Shift'
            elif PM_SHIFT_START <= current_time <= PM_SHIFT_END:
                expected_time = datetime.combine(timestamp.date(), special_expected_time_pm)
                shift = 'PM Shift'
            elif current_time < SHIFT_START:
                if special_expected_time_pm:
                    expected_time = datetime.combine(timestamp.date() - timedelta(days=1), special_expected_time_pm)
                    shift = 'PM Shift (after midnight)'
                else:
                    expected_time = None
                    shift = 'No PM Shift'
            else:
                expected_time = None
                shift = 'Unknown'

        # Check if the user is late or on time
        if expected_time:
            expected_time = LOCAL_TIME_ZONE.localize(expected_time)
            if timestamp > expected_time:
                status = 'Late'
                lateness_duration_td = timestamp - expected_time
                lateness_minutes = int(lateness_duration_td.total_seconds() // 60)
                lateness_hours = lateness_minutes // 60
                lateness_remaining_minutes = lateness_minutes % 60

                if lateness_hours > 0:
                    if lateness_remaining_minutes > 0:
                        lateness_duration = f'{lateness_hours} hrs & {lateness_remaining_minutes} mins'
                    else:
                        lateness_duration = f'{lateness_hours} hrs'
                else:
                    lateness_duration = f'{lateness_remaining_minutes} mins'
            else:
                status = 'On Time'
                lateness_duration = ''
        else:
            status = 'Invalid Time-In'
            lateness_duration = ''
            shift = ''

        # Initialize log ID
        if os.path.isfile(LOG_FILE):
            try:
                with open(LOG_FILE, 'r', newline='', encoding='utf-8') as csvfile:
                    csvreader = csv.reader(csvfile)
                    next(csvreader, None)  # Skip header
                    last_id = 0
                    for row in csvreader:
                        if row and row[0].isdigit():
                            last_id = int(row[0])
                    new_id = last_id + 1
            except Exception as e:
                app.logger.error(f"Error reading log file: {e}")
                flash('Failed to read attendance log.', 'danger')
                return redirect(url_for('index'))
        else:
            new_id = 1

        # Log the data
        data = {
            'ID': new_id,
            'Employee ID': int(employee_id),
            'Name': name,
            'Group': group.upper(),
            'Action': 'Time_In',
            'Date': date_str,
            'Start Time': time_str,
            'End Time': '',
            'Time Consumed': '',
            'Shift': shift,
            'Lateness Duration': lateness_duration,
            'Status': status
        }

        file_exists = os.path.isfile(LOG_FILE)

        try:
            with open(LOG_FILE, 'a', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['ID', 'Employee ID', 'Name', 'Group', 'Action', 'Date', 'Start Time', 'End Time', 'Time Consumed', 'Shift', 'Lateness Duration', 'Status']
                csvwriter = csv.DictWriter(csvfile, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
                if not file_exists:
                    csvwriter.writeheader()
                csvwriter.writerow(data)
        except Exception as e:
            app.logger.error(f"Error writing to log file: {e}")
            flash('Failed to record attendance. Please try again.', 'danger')
            return redirect(url_for('index'))

        # Flash appropriate message
        if status == 'Invalid Time-In':
            flash_msg = f"{status}. Please clock in during your shift hours."
        elif status == 'Overtime':
            flash_msg = f"You are {lateness_duration} late. Time-In recorded for {name} on {date_str} at {time_str}."
        else:
            flash_msg = f"{status}! Time-In recorded for {name} on {date_str} at {time_str}."

        flash(flash_msg, 'info')
        return redirect(url_for('index'))

    elif action.lower() in ['halfday_time_in', 'halfday_time_out']:
        if action.lower() == 'halfday_time_in':
            # Handle Halfday Time-In
            return handle_halfday_time_in(employee_id, name, group, timestamp, date_str, time_str)
        elif action.lower() == 'halfday_time_out':
            # Handle Halfday Time-Out
            return handle_halfday_time_out(employee_id, name, group, timestamp, date_str, time_str)

    elif action.lower() == 'time_out':
        # Check if user has a Time-In entry without End Time
        if os.path.isfile(LOG_FILE):
            try:
                df = pd.read_csv(LOG_FILE, encoding='utf-8')
                # Fill NaN values
                df['Action'] = df['Action'].fillna('')
                df['End Time'] = df['End Time'].fillna('')
                df['Date'] = df['Date'].fillna('')
                df['Start Time'] = df['Start Time'].fillna('')
                # Find the last Time-In entry for this user and date where 'End Time' is empty
                mask = (
                    (df['Employee ID'] == int(employee_id)) &
                    (df['Date'].str.strip() == date_str) &
                    (df['Action'].str.lower() == 'time_in') &
                    (df['End Time'].str.strip() == '')
                )
                if not mask.any():
                    flash('Cannot clock out without clocking in first.', 'warning')
                    return redirect(url_for('index'))
                else:
                    # Get index of the entry
                    idx = df[mask].index[-1]
                    # Update the entry
                    end_time_str = time_str
                    start_time_str = df.loc[idx, 'Start Time']
                    if not start_time_str:
                        flash('Start Time is missing for Time-In. Cannot record Time-Out.', 'danger')
                        return redirect(url_for('index'))
                    # Parse times
                    start_time = datetime.strptime(start_time_str, '%H:%M:%S')
                    end_time = datetime.strptime(end_time_str, '%H:%M:%S')
                    # If end_time < start_time, it means the end time is on the next day
                    if end_time < start_time:
                        end_time += timedelta(days=1)
                    duration_td = end_time - start_time
                    duration_seconds = duration_td.total_seconds()

                    # Extract hours, minutes, and seconds
                    duration_hours = int(duration_seconds // 3600)
                    duration_remaining_seconds = int(duration_seconds % 3600)
                    duration_minutes = duration_remaining_seconds // 60
                    duration_secs = duration_remaining_seconds % 60

                    # Format Time Consumed
                    duration_str_parts = []
                    if duration_hours > 0:
                        duration_str_parts.append(f"{duration_hours} hrs")
                    if duration_minutes > 0:
                        duration_str_parts.append(f"{duration_minutes} mins")
                    if duration_secs > 0:
                        duration_str_parts.append(f"{duration_secs} secs")
                    duration_str = ' & '.join(duration_str_parts) if duration_str_parts else '0 secs'
                    df.loc[idx, 'End Time'] = end_time_str
                    df.loc[idx, 'Time Consumed'] = duration_str
                    df.loc[idx, 'Action'] = 'Time_in/Time_out'
                    # Save the DataFrame back to CSV
                    df.to_csv(LOG_FILE, index=False, quoting=csv.QUOTE_ALL)
                    flash(f"Time-Out recorded for {name} on {date_str} at {end_time_str}.", 'info')
                    return redirect(url_for('index'))
            except Exception as e:
                app.logger.error(f"Error processing Time-Out: {e}")
                flash('Failed to record Time-Out. Please try again.', 'danger')
                return redirect(url_for('index'))
        else:
            flash('Cannot clock out without clocking in first.', 'warning')
            return redirect(url_for('index'))

    elif action in TIME_LIMITS:
        # Generate a new log ID
        if os.path.isfile(LOG_FILE):
            try:
                with open(LOG_FILE, 'r', newline='', encoding='utf-8') as csvfile:
                    csvreader = csv.reader(csvfile)
                    next(csvreader, None)  # Skip header
                    last_id = 0
                    for row in csvreader:
                        if row and row[0].isdigit():
                            last_id = int(row[0])
                    new_id = last_id + 1
            except Exception as e:
                app.logger.error(f"Error reading log file: {e}")
                flash('Failed to record action. Please try again.', 'danger')
                return redirect(url_for('index'))
        else:
            new_id = 1

        timestamp = get_pakistan_time()
        date_str = timestamp.strftime('%Y-%m-%d')
        start_time_str = timestamp.strftime('%H:%M:%S')

        # Log the action immediately with Start Time
        data = {
            'ID': new_id,
            'Employee ID': int(employee_id),
            'Name': name,
            'Group': group.upper(),
            'Action': action,
            'Date': date_str,
            'Start Time': start_time_str,
            'End Time': '',
            'Time Consumed': '',
            'Shift': '',
            'Lateness Duration': '',
            'Status': ''
        }

        file_exists = os.path.isfile(LOG_FILE)

        try:
            with open(LOG_FILE, 'a', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['ID', 'Employee ID', 'Name', 'Group', 'Action', 'Date', 'Start Time', 'End Time', 'Time Consumed', 'Shift', 'Lateness Duration', 'Status']
                csvwriter = csv.DictWriter(csvfile, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
                if not file_exists:
                    csvwriter.writeheader()
                csvwriter.writerow(data)
        except Exception as e:
            app.logger.error(f"Error writing to log file: {e}")
            flash('Failed to record action. Please try again.', 'danger')
            return redirect(url_for('index'))

        # Store the log ID and other data in the temp file
        identifier = str(uuid.uuid4())

        # Ensure temp directory exists
        TEMP_DIR = os.path.join(BASE_DIR, 'temp')
        if not os.path.exists(TEMP_DIR):
            os.makedirs(TEMP_DIR)

        temp_data = {
            'identifier': identifier,
            'log_id': new_id,
            'employee_id': employee_id,
            'name': name,
            'group': group,
            'action': action,
            'start_time': timestamp.strftime('%Y-%m-%d %H:%M:%S')
        }
        temp_file_path = os.path.join(TEMP_DIR, f"{identifier}.json")
        try:
            with open(temp_file_path, 'w') as temp_file:
                json.dump(temp_data, temp_file)
        except Exception as e:
            app.logger.error(f"Error writing temp file: {e}")
            flash('Failed to start action. Please try again.', 'danger')
            return redirect(url_for('index'))

        # Return a page with 'Back to Work' button
        return render_template('back_to_work.html', identifier=identifier, action=action)

    else:
        flash('Invalid action selected.', 'danger')
        return redirect(url_for('index'))


def handle_break_action(employee_id, name, group, action, timestamp, date_str, time_str):
    """
    Handles break actions by recording them without duplication restrictions.
    """
    # Initialize log ID
    new_id = get_next_log_id()

    # Log the data
    data = {
        'ID': new_id,
        'Employee ID': int(employee_id),
        'Name': name,
        'Group': group.upper(),
        'Action': action,
        'Date': date_str,
        'Start Time': time_str,
        'End Time': '',
        'Time Consumed': '',
        'Shift': '',
        'Lateness Duration': '',
        'Status': ''
    }

    # Append to log.csv
    append_to_log_file(data)

    flash(f"Action '{action}' recorded for {name} on {date_str} at {time_str}.", 'info')
    return redirect(url_for('index'))


# New route to handle 'Back to Work'
@app.route('/attendance/back_to_work', methods=['POST'])
def back_to_work():
    identifier = request.form.get('identifier', '').strip()
    if not identifier:
        flash('Invalid request. Missing identifier.', 'danger')
        return redirect(url_for('index'))

    # Read the temp file
    TEMP_DIR = os.path.join(BASE_DIR, 'temp')
    temp_file_path = os.path.join(TEMP_DIR, f"{identifier}.json")
    if not os.path.exists(temp_file_path):
        flash('Session expired or invalid identifier.', 'danger')
        return redirect(url_for('index'))

    try:
        with open(temp_file_path, 'r') as temp_file:
            temp_data = json.load(temp_file)
    except Exception as e:
        app.logger.error(f"Error reading temp file: {e}")
        flash('Failed to retrieve action data.', 'danger')
        return redirect(url_for('index'))

    # Get the data
    log_id = temp_data['log_id']
    employee_id = temp_data['employee_id']
    name = temp_data['name']
    group = temp_data['group']
    action = temp_data['action']
    start_time_str = temp_data['start_time']

    # Parse start time
    start_time = datetime.strptime(start_time_str, '%Y-%m-%d %H:%M:%S')
    start_time = LOCAL_TIME_ZONE.localize(start_time)

    # Get end time
    end_time = get_pakistan_time()

    # Calculate duration
    duration_td = end_time - start_time
    duration_seconds = duration_td.total_seconds()

    # Extract hours, minutes, and seconds
    duration_hours = int(duration_seconds // 3600)
    duration_remaining_seconds = int(duration_seconds % 3600)
    duration_minutes = duration_remaining_seconds // 60
    duration_secs = duration_remaining_seconds % 60

    # Compare with time limit
    time_limit = TIME_LIMITS.get(action, 0)  # Time limit in minutes
    time_limit_seconds = time_limit * 60

    if duration_seconds <= time_limit_seconds:
        status = 'On Time'
        lateness_duration = ''
    else:
        status = 'Overbreak'
        overbreak_seconds = duration_seconds - time_limit_seconds
        # Calculate overbreak hours, minutes, and seconds
        lateness_hours = int(overbreak_seconds // 3600)
        lateness_remaining_seconds = int(overbreak_seconds % 3600)
        lateness_minutes = lateness_remaining_seconds // 60
        lateness_secs = lateness_remaining_seconds % 60

        # Format lateness_duration
        parts = []
        if lateness_hours > 0:
            parts.append(f"{lateness_hours} hrs")
        if lateness_minutes > 0:
            parts.append(f"{lateness_minutes} mins")
        if lateness_secs > 0:
            parts.append(f"{lateness_secs} secs")
        lateness_duration = ' & '.join(parts)

    # Prepare data to log
    date_str = start_time.strftime('%Y-%m-%d')
    end_time_str = end_time.strftime('%H:%M:%S')
    duration_str_parts = []
    if duration_hours > 0:
        duration_str_parts.append(f"{duration_hours} hrs")
    if duration_minutes > 0:
        duration_str_parts.append(f"{duration_minutes} mins")
    if duration_secs > 0:
        duration_str_parts.append(f"{duration_secs} secs")
    duration_str = ' & '.join(duration_str_parts) if duration_str_parts else '0 secs'

    # Initialize log ID
    if os.path.isfile(LOG_FILE):
        try:
            with open(LOG_FILE, 'r', newline='', encoding='utf-8') as csvfile:
                csvreader = csv.reader(csvfile)
                next(csvreader, None)  # Skip header
                last_id = 0
                for row in csvreader:
                    if row and row[0].isdigit():
                        last_id = int(row[0])
                # Ensure unique ID
                new_id = last_id + 1
        except Exception as e:
            app.logger.error(f"Error reading log file: {e}")
            flash('Failed to read attendance log.', 'danger')
            return redirect(url_for('index'))
    else:
        new_id = 1

    # Read the log CSV into a DataFrame
    try:
        df = pd.read_csv(LOG_FILE, encoding='utf-8')
    except Exception as e:
        app.logger.error(f"Error reading log file: {e}")
        flash('Failed to read attendance log.', 'danger')
        return redirect(url_for('index'))

    # Find the log entry by ID
    mask = df['ID'] == log_id
    if not mask.any():
        flash('Log entry not found. Cannot update.', 'danger')
        return redirect(url_for('index'))

    # Update the entry
    df.loc[mask, 'End Time'] = end_time_str
    df.loc[mask, 'Time Consumed'] = duration_str
    df.loc[mask, 'Lateness Duration'] = lateness_duration
    df.loc[mask, 'Status'] = status

    # Save the DataFrame back to CSV
    try:
        df.to_csv(LOG_FILE, index=False, quoting=csv.QUOTE_ALL)
    except Exception as e:
        app.logger.error(f"Error updating log file: {e}")
        flash('Failed to update action. Please try again.', 'danger')
        return redirect(url_for('index'))

    # Delete the temp file
    try:
        os.remove(temp_file_path)
    except Exception as e:
        app.logger.error(f"Error deleting temp file: {e}")

    flash(f"Action '{action}' completed. You spent {duration_str}. Status: {status}.", 'info')
    return redirect(url_for('index'))



@app.route('/attendance/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        input_key = request.form.get('master_key', '').strip()
        stored_master_key, stored_sub_keys = get_keys()

        if not stored_master_key:
            flash('Master key not set. Please contact the administrator.', 'danger')
            return redirect(url_for('login'))

        if input_key == stored_master_key:
            session['authenticated'] = True
            session['role'] = 'admin'  # Full access
            flash('Logged in as admin.', 'success')
            return redirect(url_for('report'))
        elif input_key in stored_sub_keys:
            session['authenticated'] = True
            session['role'] = 'user'  # Restricted access
            flash('Logged in as user.', 'success')
            return redirect(url_for('report'))
        else:
            flash('Invalid key. Access denied.', 'danger')
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/attendance/add_sub_key', methods=['GET', 'POST'])
@login_required
@admin_required
def add_sub_key():
    if request.method == 'POST':
        new_sub_key = request.form.get('new_sub_key', '').strip()
        if not new_sub_key:
            flash('Sub-key cannot be empty.', 'warning')
            return redirect(url_for('add_sub_key'))

        master_key, sub_keys = get_keys()

        if new_sub_key in sub_keys:
            flash('Sub-key already exists.', 'warning')
        else:
            sub_keys.append(new_sub_key)
            set_keys(master_key, sub_keys)
            flash('Sub-key added successfully.', 'success')

        return redirect(url_for('manage_sub_keys'))

    return render_template('add_sub_key.html')

@app.route('/attendance/delete_sub_key', methods=['GET', 'POST'])
@login_required
@admin_required
def delete_sub_key():
    master_key, sub_keys = get_keys()

    if request.method == 'POST':
        sub_key_to_delete = request.form.get('sub_key_to_delete', '').strip()
        if sub_key_to_delete in sub_keys:
            sub_keys.remove(sub_key_to_delete)
            set_keys(master_key, sub_keys)
            flash('Sub-key deleted successfully.', 'success')
        else:
            flash('Sub-key not found.', 'warning')
        return redirect(url_for('manage_sub_keys'))

    return render_template('delete_sub_key.html', sub_keys=sub_keys)

@app.route('/attendance/change_sub_key', methods=['GET', 'POST'])
@login_required
@admin_required
def change_sub_key():
    master_key, sub_keys = get_keys()

    if request.method == 'POST':
        old_sub_key = request.form.get('old_sub_key', '').strip()
        new_sub_key = request.form.get('new_sub_key', '').strip()

        if not old_sub_key or not new_sub_key:
            flash('Both old and new sub-keys are required.', 'warning')
            return redirect(url_for('change_sub_key'))

        if old_sub_key in sub_keys:
            if new_sub_key in sub_keys:
                flash('The new sub-key already exists.', 'warning')
            else:
                sub_keys[sub_keys.index(old_sub_key)] = new_sub_key
                set_keys(master_key, sub_keys)
                flash('Sub-key changed successfully.', 'success')
        else:
            flash('Old sub-key not found.', 'warning')

        return redirect(url_for('manage_sub_keys'))

    return render_template('change_sub_key.html', sub_keys=sub_keys)

@app.route('/attendance/manage_sub_keys')
@login_required
@admin_required
def manage_sub_keys():
    master_key, sub_keys = get_keys()
    return render_template('manage_sub_keys.html', sub_keys=sub_keys)

@app.route('/attendance/report')
@login_required
def report():
    if os.path.isfile(LOG_FILE):
        try:
            df = pd.read_csv(LOG_FILE, encoding='utf-8')
            df = df.fillna('')  # Replace NaN with empty string
            df = df.sort_values(by='ID', ascending=False)
            data = df.values.tolist()
            headers = df.columns.tolist()

            # Identify the index of the 'Action' column
            try:
                action_idx = headers.index('Action')
            except ValueError:
                app.logger.error("'Action' column not found in the log file.")
                flash("'Action' column is missing from the log data.", 'danger')
                data = []
                headers = []
                return render_template('report.html', data=data, headers=headers)

            # Replace 'Halfday_Time_In' and 'Halfday_Time_Out' with 'Halfday_Time_In/Halfday_Time_Out'
            for row in data:
                if row[action_idx] in ['Halfday_Time_In', 'Halfday_Time_Out']:
                    row[action_idx] = 'Halfday_Time_In/Halfday_Time_Out'

        except Exception as e:
            app.logger.error(f"Error reading log file: {e}")
            flash('Failed to load attendance data.', 'danger')
            data = []
            headers = []
    else:
        data = []
        headers = []

    return render_template('report.html', data=data, headers=headers)


@app.route('/attendance/export')
@login_required
def export():
    if os.path.isfile(LOG_FILE):
        try:
            # Read the log CSV into a DataFrame
            df = pd.read_csv(LOG_FILE, encoding='utf-8')
            df = df.fillna('')  # Replace NaN with empty string
            df = df.sort_values(by='ID', ascending=False)

            # Define break actions
            BREAK_ACTIONS = ["Recite Sutra", "Toilet", "Smoke", "BREAK1", "BREAK2"]
            HALFDAY_ACTIONS = ["Halfday_Time_In", "Halfday_Time_Out"]

            # Separate regular attendance and breaks
            attendance_df = df[~df['Action'].isin(BREAK_ACTIONS + HALFDAY_ACTIONS)]
            breaks_df = df[df['Action'].isin(BREAK_ACTIONS)]
            halfday_df = df[df['Action'].isin(HALFDAY_ACTIONS)]

            # Create a BytesIO buffer to hold the Excel file in memory
            output = BytesIO()

            # Use ExcelWriter with openpyxl engine
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Write regular attendance to 'Attendance' sheet
                attendance_df.to_excel(writer, index=False, sheet_name='Attendance')

                # Write breaks to 'Breaks' sheet
                breaks_df.to_excel(writer, index=False, sheet_name='Breaks')

                # Write half-day actions to 'Halfday' sheet
                halfday_df.to_excel(writer, index=False, sheet_name='Halfday')

                # Access the workbook and worksheets
                workbook = writer.book
                attendance_sheet = writer.sheets['Attendance']
                breaks_sheet = writer.sheets['Breaks']

                # Define the fills for highlighting
                late_fill = PatternFill(start_color='FDEF81', end_color='FDEF81', fill_type='solid')  # Light yellow fill
                overbreak_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')  # Light red fill

                # Function to apply conditional formatting
                def apply_conditional_formatting(sheet, status_column):
                    for row_idx, status in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row), start=2):
                        cell_status = sheet.cell(row=row_idx, column=sheet.max_row).value  # Assuming 'Status' is the last column
                        fill = None
                        if str(status).strip().lower() == 'late':
                            fill = late_fill
                        elif str(status).strip().lower() == 'overbreak':
                            fill = overbreak_fill

                        if fill:
                            for col in range(1, sheet.max_column + 1):
                                cell = sheet.cell(row=row_idx, column=col)
                                cell.fill = fill

                # Apply formatting to 'Attendance' sheet
                if 'Status' in attendance_df.columns:
                    apply_conditional_formatting(attendance_sheet, 'Status')
                else:
                    app.logger.error("'Status' column not found in the Attendance sheet.")
                    flash("'Status' column is missing from the Attendance data.", 'danger')
                    return redirect(url_for('report'))

                # Apply formatting to 'Breaks' sheet (if needed)
                if 'Status' in breaks_df.columns:
                    apply_conditional_formatting(breaks_sheet, 'Status')
                else:
                    app.logger.error("'Status' column not found in the Breaks sheet.")
                    flash("'Status' column is missing from the Breaks data.", 'danger')
                    return redirect(url_for('report'))

            # Seek to the beginning of the BytesIO buffer
            output.seek(0)

            # Generate a dynamic filename with the current date and time
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'attendance_report_{timestamp}.xlsx'

            return send_file(
                output,
                download_name=filename,
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            app.logger.error(f"Error exporting to Excel: {e}")
            flash('Failed to export attendance data.', 'danger')
    else:
        flash('No attendance data available to export.', 'warning')

    return redirect(url_for('report'))


@app.route('/attendance/logout')
@login_required
def logout():
    session.pop('authenticated', None)
    session.pop('role', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/attendance/change_key', methods=['GET', 'POST'])
@login_required
@admin_required
def change_key():
    if request.method == 'POST':
        new_master_key = request.form.get('new_master_key', '').strip()

        if not new_master_key:
            flash('Master key is required.', 'warning')
            return redirect(url_for('change_key'))

        master_key, sub_keys = get_keys()
        set_keys(new_master_key, sub_keys)
        flash('Master key updated successfully.', 'success')
        return redirect(url_for('report'))

    return render_template('change_key.html')

@app.route('/attendance/purge_duplicates', methods=['GET', 'POST'])
@login_required
@admin_required
def purge_duplicates():
    if request.method == 'POST':
        success, message = purge_duplicate_actions()
        if success:
            flash(message, 'success')
        else:
            flash(message, 'warning')
        return redirect(url_for('report'))
    return render_template('confirm_purge.html')

@app.route('/attendance/manage_employees')
@login_required
@admin_required
def manage_employees():
    employee_list = get_employee_list()
    return render_template('manage_employees.html', employee_list=employee_list)

@app.route('/attendance/add_employee', methods=['GET', 'POST'])
@login_required
@admin_required
def add_employee():
    if request.method == 'POST':
        employee_name = request.form.get('employee_name', '').strip()
        if not employee_name:
            flash('Employee Name is required.', 'warning')
            return redirect(url_for('add_employee'))

        # Read existing employees
        employee_list = get_employee_list()

        # Determine the next employee ID with leading zeros
        if employee_list:
            try:
                # Extract existing IDs as integers
                existing_ids = [int(emp['ID']) for emp in employee_list]
                max_id = max(existing_ids)
                new_id_int = max_id + 1
                # Format the new_id as a four-digit string with leading zeros
                new_id = f"{new_id_int:04}"
            except ValueError:
                flash('Existing employee IDs are not numeric. Cannot auto-increment.', 'danger')
                return redirect(url_for('add_employee'))
        else:
            new_id = '0001'  # Start IDs at '0001' if no employees exist

        # Append the new employee to employees.csv
        try:
            with open(EMPLOYEES_FILE, 'a', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['ID', 'Name']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                # Write header if the file is empty
                if os.stat(EMPLOYEES_FILE).st_size == 0:
                    writer.writeheader()
                writer.writerow({'ID': new_id, 'Name': employee_name})
            flash(f'Employee "{employee_name}" added successfully with ID {new_id}.', 'success')
        except Exception as e:
            app.logger.error(f"Error adding employee: {e}")
            flash('Failed to add employee.', 'danger')
        return redirect(url_for('manage_employees'))

    # For GET requests, determine the next available ID with leading zeros
    employee_list = get_employee_list()
    if employee_list:
        try:
            existing_ids = [int(emp['ID']) for emp in employee_list]
            max_id = max(existing_ids)
            next_id_int = max_id + 1
            next_id = f"{next_id_int:04}"
        except ValueError:
            next_id = 'N/A'
            flash('Existing employee IDs are not numeric. Cannot auto-increment.', 'danger')
    else:
        next_id = '0001'

    return render_template('add_employee.html', next_id=next_id)


@app.route('/attendance/edit_employee/<employee_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_employee(employee_id):
    employee_list = get_employee_list()
    employee = next((emp for emp in employee_list if emp['ID'] == employee_id), None)
    if not employee:
        flash('Employee not found.', 'danger')
        return redirect(url_for('manage_employees'))

    if request.method == 'POST':
        new_employee_name = request.form.get('employee_name', '').strip()
        if not new_employee_name:
            flash('Employee Name is required.', 'warning')
            return redirect(url_for('edit_employee', employee_id=employee_id))

        # Update the employee in the list
        for emp in employee_list:
            if emp['ID'] == employee_id:
                emp['Name'] = new_employee_name
                break
        # Write the updated list back to employees.csv
        try:
            with open(EMPLOYEES_FILE, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['ID', 'Name']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(employee_list)
            flash('Employee updated successfully.', 'success')
        except Exception as e:
            app.logger.error(f"Error updating employee: {e}")
            flash('Failed to update employee.', 'danger')
        return redirect(url_for('manage_employees'))
    return render_template('edit_employee.html', employee=employee)

@app.route('/attendance/delete_employee/<employee_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def delete_employee(employee_id):
    employee_list = get_employee_list()
    employee = next((emp for emp in employee_list if emp['ID'] == employee_id), None)
    if not employee:
        flash('Employee not found.', 'danger')
        return redirect(url_for('manage_employees'))

    if request.method == 'POST':
        # Remove the employee from the list
        employee_list = [emp for emp in employee_list if emp['ID'] != employee_id]
        # Write the updated list back to employees.csv
        try:
            with open(EMPLOYEES_FILE, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['ID', 'Name']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(employee_list)
            flash('Employee deleted successfully.', 'success')
        except Exception as e:
            app.logger.error(f"Error deleting employee: {e}")
            flash('Failed to delete employee.', 'danger')
        return redirect(url_for('manage_employees'))
    return render_template('delete_employee.html', employee=employee)

@app.errorhandler(403)
def forbidden(e):
    return render_template('403.html'), 403

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    app.logger.error(f"Server Error: {e}")
    return render_template('500.html'), 500

if __name__ == '__main__':
    # It's recommended to set debug to False in production
    app.run(debug=True, host='0.0.0.0', port=8003)
